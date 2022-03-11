import openpyxl

def store(dcp_psnr,dcp_ssim,erc_psnr,erc_ssim,img):
    wb = openpyxl.Workbook()
    sheet = wb.active
    c1 = sheet.cell(row = 1, column = 1)
    c1.value = "HAZY IMAGE"
    c1 = sheet.cell(row = 1, column = 2)
    c1.value = "dcp-SSIM"
    c2 = sheet.cell(row = 1, column = 3)
    c2.value = "erc-SSIM"
    c1 = sheet.cell(row = 1, column = 4)
    c1.value = "dcp-PSNR"
    c2 = sheet.cell(row = 1, column = 5)
    c2.value = "erc-PSNR"
    for i in range(len(dcp_ssim)):
        c1 = sheet.cell(row = i+2, column = 2)
        c1.value = dcp_ssim[i]
    for i in range(len(erc_ssim)):
        c2 = sheet.cell(row = i+2, column = 3)
        c2.value = erc_ssim[i]
    for i in range(len(img)):
        c1 = sheet.cell(row = i+2, column = 1)
        c1.value = img[i]
    for i in range(len(dcp_psnr)):
        c1 = sheet.cell(row = i+2, column = 4)
        c1.value = dcp_psnr[i]
    for i in range(len(erc_psnr)):
        c2 = sheet.cell(row = i+2, column = 5)
        c2.value = erc_psnr[i]
    wb.save("sample.xlsx") 

